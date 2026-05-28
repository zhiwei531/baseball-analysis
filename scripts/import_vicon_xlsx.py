from __future__ import annotations

import argparse
from pathlib import Path
import sys
from typing import Iterable

from openpyxl import load_workbook

PROJECT_SRC = Path(__file__).resolve().parents[1] / "src"
if str(PROJECT_SRC) not in sys.path:
    sys.path.insert(0, str(PROJECT_SRC))

from baseball_pose.io.pose_csv import write_pose_records
from baseball_pose.pose.schema import PoseRecord


MARKER_GROUPS = {
    "nose": ("LFHD", "RFHD", "LBHD", "RBHD"),
    "left_shoulder": ("LSHO",),
    "right_shoulder": ("RSHO",),
    "left_elbow": ("LELB",),
    "right_elbow": ("RELB",),
    "left_wrist": ("LWRA", "LWRB"),
    "right_wrist": ("RWRA", "RWRB"),
    "left_hip": ("LASI", "LPSI"),
    "right_hip": ("RASI", "RPSI"),
    "left_knee": ("LKNE",),
    "right_knee": ("RKNE",),
    "left_ankle": ("LANK",),
    "right_ankle": ("RANK",),
}

AXIS_INDEX = {"x": 0, "y": 1, "z": 2}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert a Vicon trajectory XLSX into baseball_pose pose CSV.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--clip-id", required=True)
    parser.add_argument("--condition-id", default="vicon_xy_projected")
    parser.add_argument("--out", required=True)
    parser.add_argument("--horizontal-axis", choices=("x", "y", "z"), default="x")
    parser.add_argument("--vertical-axis", choices=("x", "y", "z"), default="y")
    parser.add_argument("--backend", default="vicon_xlsx_projection")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    output_path = Path(args.out).resolve()

    sample_rate_hz, frame_rows = load_vicon_rows(xlsx_path)
    records = build_pose_records(
        clip_id=args.clip_id,
        condition_id=args.condition_id,
        frame_rows=frame_rows,
        sample_rate_hz=sample_rate_hz,
        horizontal_axis=args.horizontal_axis,
        vertical_axis=args.vertical_axis,
        backend=args.backend,
    )
    write_pose_records(output_path, records)
    print(output_path)
    print(f"sample_rate_hz={sample_rate_hz}")
    print(f"frames={len(frame_rows)}")
    print(f"records={len(records)}")


def load_vicon_rows(path: Path) -> tuple[float, list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    sample_rate_hz = float(sheet["A2"].value or 100.0)
    marker_header = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))

    columns: list[tuple[str, int]] = []
    for idx, value in enumerate(marker_header):
        if idx < 2 or value is None:
            continue
        marker_name = str(value).split(":")[-1]
        columns.append((marker_name, idx))

    frame_rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=6, values_only=True):
        frame_value = row[0]
        if frame_value is None:
            continue
        marker_values: dict[str, tuple[float | None, float | None, float | None]] = {}
        for marker_name, start_idx in columns:
            xyz = row[start_idx : start_idx + 3]
            if len(xyz) < 3:
                continue
            marker_values[marker_name] = (
                optional_float(xyz[0]),
                optional_float(xyz[1]),
                optional_float(xyz[2]),
            )
        frame_rows.append(
            {
                "frame_index": int(frame_value) - 1,
                "markers": marker_values,
            }
        )
    return sample_rate_hz, frame_rows


def build_pose_records(
    clip_id: str,
    condition_id: str,
    frame_rows: list[dict[str, object]],
    sample_rate_hz: float,
    horizontal_axis: str,
    vertical_axis: str,
    backend: str,
) -> list[PoseRecord]:
    axis_x = AXIS_INDEX[horizontal_axis]
    axis_y = AXIS_INDEX[vertical_axis]

    frame_points = []
    raw_xy: list[tuple[float, float]] = []
    for row in frame_rows:
        markers = row["markers"]
        assert isinstance(markers, dict)
        joints: dict[str, tuple[float, float, float] | None] = {}
        for joint_name, marker_names in MARKER_GROUPS.items():
            joints[joint_name] = mean_marker(markers, marker_names)
            point = joints[joint_name]
            if point is not None:
                raw_xy.append((point[axis_x], point[axis_y]))
        frame_points.append((int(row["frame_index"]), joints))

    min_x = min(value[0] for value in raw_xy)
    max_x = max(value[0] for value in raw_xy)
    min_y = min(value[1] for value in raw_xy)
    max_y = max(value[1] for value in raw_xy)
    width = max(max_x - min_x, 1.0)
    height = max(max_y - min_y, 1.0)

    records: list[PoseRecord] = []
    for frame_index, joints in frame_points:
        timestamp_sec = frame_index / sample_rate_hz
        for joint_name, point in joints.items():
            x_norm: float | None
            y_norm: float | None
            visibility: float | None
            confidence: float | None
            if point is None:
                x_norm = None
                y_norm = None
                visibility = None
                confidence = None
            else:
                x_norm = (point[axis_x] - min_x) / width
                y_norm = 1.0 - ((point[axis_y] - min_y) / height)
                visibility = 1.0
                confidence = 1.0
            records.append(
                PoseRecord(
                    clip_id=clip_id,
                    condition_id=condition_id,
                    frame_index=frame_index,
                    timestamp_sec=timestamp_sec,
                    joint_name=joint_name,
                    x=x_norm,
                    y=y_norm,
                    visibility=visibility,
                    confidence=confidence,
                    backend=backend,
                    inference_time_ms=None,
                )
            )
    return records


def mean_marker(
    markers: dict[str, tuple[float | None, float | None, float | None]],
    names: Iterable[str],
) -> tuple[float, float, float] | None:
    valid = [markers[name] for name in names if name in markers and all(value is not None for value in markers[name])]
    if not valid:
        return None
    count = float(len(valid))
    return (
        sum(float(item[0]) for item in valid) / count,
        sum(float(item[1]) for item in valid) / count,
        sum(float(item[2]) for item in valid) / count,
    )


def optional_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


if __name__ == "__main__":
    main()
